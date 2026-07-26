"""Agent 工具集：时间查询 + 联网搜索 + 知识库检索（RAG）+ 业务系统集成"""

import logging
import threading
from datetime import datetime

from langchain_core.tools import tool
from tavily import TavilyClient

from .config import get_settings
from .retry import log_execution, retry_on_error
from .sanitizers import sanitize_exception
from .validators import InputValidationError, validate_search_query

logger = logging.getLogger(__name__)


# P1 性能：Tavily client 单例化，避免每次 web_search 都新建 client（含内部 httpx client）
# TavilyClient 内部维护 httpx.Client，复用可减少 TCP+TLS 握手开销
_tavily_client: TavilyClient | None = None
_tavily_client_lock = threading.Lock()


def get_tavily_client() -> TavilyClient | None:
    """获取 Tavily client 单例（线程安全）。

    未配置 TAVILY_API_KEY 时返回 None。
    """
    global _tavily_client
    if _tavily_client is None:
        with _tavily_client_lock:
            if _tavily_client is None:
                api_key = get_settings().tavily_api_key
                if not api_key:
                    return None
                _tavily_client = TavilyClient(api_key=api_key)
    return _tavily_client


def reset_tavily_client() -> None:
    """重置 Tavily client 单例（主要用于测试与配置切换）"""
    global _tavily_client
    with _tavily_client_lock:
        _tavily_client = None


# 1. 基础工具：获取时间
@tool
@log_execution
def get_current_time():
    """获取当前的日期和时间。当用户询问现在几点、今天日期、星期几时使用此工具。"""
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# 2. 联网搜索工具（使用 Tavily）
@tool
@log_execution
def web_search(query: str):
    """在真实互联网上搜索高质量信息。当用户询问最新的新闻、天气、股票、体育赛事或任何需要实时外部知识的问题时，必须使用此工具。"""
    try:
        # 输入验证
        query = validate_search_query(query)
    except InputValidationError as e:
        logger.warning("搜索查询验证失败: %s", e)
        return f"❌ 输入无效: {e}"

    try:
        # P1 性能：复用 Tavily client 单例，避免每次新建 client + httpx.Client
        client = get_tavily_client()
        if client is None:
            logger.warning("TAVILY_API_KEY 未配置，联网搜索不可用")
            return "❌ 错误：未找到 TAVILY_API_KEY，请检查项目根目录下的 .env 文件配置。"

        logger.info("执行联网搜索: %s", query)

        # 带重试的搜索调用
        response = _search_with_retry(client, query)

        if not response.get("results"):
            logger.info("搜索无结果: %s", query)
            return "抱歉，未能通过 Tavily 检索到相关信息。"

        # 格式化输出
        formatted_results = []
        for r in response["results"]:
            formatted_results.append(f"标题: {r['title']}\n" f"摘要: {r['content']}\n" f"链接: {r['url']}")

        search_result = "\n\n".join(formatted_results)
        logger.info("搜索完成，返回 %d 条结果", len(response["results"]))
        return f"搜索结果:\n{search_result}"

    except Exception as e:
        logger.exception("联网搜索发生错误")
        return f"联网搜索发生错误: {sanitize_exception(e)}"


@retry_on_error(max_attempts=3, min_wait=1.0, max_wait=8.0)
def _search_with_retry(client: TavilyClient, query: str) -> dict:
    """带指数退避重试的 Tavily 搜索"""
    return client.search(query, max_results=3)


# 3. 知识库检索工具（RAG）
@tool
@log_execution
def knowledge_search(query: str):
    """
    从本地知识库中检索相关文档信息。
    当用户询问项目文档、产品手册、内部知识、历史对话等私有知识时使用此工具。
    如果知识库中没有相关信息，请提示用户使用 knowledge_upload 工具上传文档。
    """
    try:
        query = validate_search_query(query)
    except InputValidationError as e:
        logger.warning("知识库检索查询验证失败: %s", e)
        return f"❌ 输入无效: {e}"

    try:
        from .rag import get_rag_manager

        manager = get_rag_manager()
        results = manager.search(query)

        if not results:
            logger.info("知识库检索无结果: %s", query)
            return "知识库中未找到相关信息。您可以上传文档到知识库后再试。"

        # 格式化检索结果
        formatted = []
        for i, doc in enumerate(results, 1):
            source = doc.metadata.get("source", "未知来源")
            formatted.append(f"【片段 {i}】(来源: {source})\n{doc.page_content}")

        logger.info("知识库检索完成: query='%s', 返回 %d 个片段", query[:50], len(results))
        return "\n\n".join(formatted)

    except ImportError as e:
        logger.warning("RAG 依赖未安装: %s", e)
        return "❌ 知识库功能不可用：缺少依赖（chromadb / langchain-chroma）"
    except Exception as e:
        logger.exception("知识库检索发生错误")
        return f"知识库检索发生错误: {sanitize_exception(e)}"


# 4. 知识库上传工具
@tool
@log_execution
def knowledge_upload(text: str, source: str = "user_input"):
    """
    将文本上传到本地知识库，供后续检索使用。
    当用户希望让 Agent 记住某些知识、文档或信息时使用此工具。

    Args:
        text: 要上传的文本内容
        source: 文档来源标识（可选，默认为 user_input）
    """
    try:
        # P0 修复：工具内补齐输入校验与敏感检测，与 HTTP /knowledge/upload 端点保持一致
        # 防止通过 LLM 工具调用绕过 HTTP 层的输入校验
        from .api import _scan_knowledge_sensitive
        from .validators import validate_knowledge_text

        try:
            clean_text = validate_knowledge_text(text)
        except Exception as e:
            return f"❌ 输入无效: {e}"
        _scan_knowledge_sensitive(clean_text, source=source)

        from .rag import get_rag_manager

        manager = get_rag_manager()
        count = manager.add_documents(clean_text, source=source)

        if count == 0:
            return "❌ 上传失败：文本为空或切片后无内容"

        return f"✅ 上传成功：文本已切分为 {count} 个片段并存入知识库"

    except ImportError as e:
        logger.warning("RAG 依赖未安装: %s", e)
        return "❌ 知识库功能不可用：缺少依赖（chromadb / langchain-chroma）"
    except Exception as e:
        logger.exception("知识库上传发生错误")
        return f"知识库上传发生错误: {sanitize_exception(e)}"


# 5. 业务系统集成工具：CRM 客户查询
@tool
@log_execution
def crm_query_customer(customer_id: str):
    """
    从 CRM 系统查询客户信息，包括基本资料和订单汇总。
    当用户询问客户信息、客户详情、客户消费记录时使用此工具。

    Args:
        customer_id: 客户ID（如 C001、C002）
    """
    try:
        from .integrations import get_crm_client

        client = get_crm_client()
        summary = client.get_customer_summary(customer_id)

        if "error" in summary:
            return f"❌ {summary['error']}"

        c = summary["customer"]
        orders_text = (
            "\n".join(
                f"  - {o['order_id']}: {o['product']} | ¥{o['amount']:.2f} | {o['status']} | {o['date']}"
                for o in summary["recent_orders"]
            )
            or "  无订单记录"
        )

        return (
            f"📋 客户信息\n"
            f"  ID: {c['customer_id']}\n"
            f"  姓名: {c['name']}\n"
            f"  公司: {c['company']}\n"
            f"  邮箱: {c['email']}\n"
            f"  电话: {c['phone']}\n"
            f"  等级: {c['level']}\n"
            f"  状态: {c['status']}\n\n"
            f"📊 订单统计\n"
            f"  订单总数: {summary['order_count']}\n"
            f"  已完成消费: ¥{summary['total_spent']:.2f}\n\n"
            f"📦 近期订单\n{orders_text}"
        )

    except Exception as e:
        logger.exception("CRM 客户查询失败")
        return f"CRM 客户查询失败: {sanitize_exception(e)}"


# 6. 业务系统集成工具：CRM 客户搜索
@tool
@log_execution
def crm_search_customers(keyword: str):
    """
    在 CRM 系统中按关键词搜索客户（支持姓名、公司、邮箱模糊匹配）。
    当用户想查找某个公司或某个人的客户信息时使用此工具。

    Args:
        keyword: 搜索关键词（客户姓名、公司名或邮箱）
    """
    try:
        keyword = validate_search_query(keyword)
    except InputValidationError as e:
        return f"❌ 输入无效: {e}"

    try:
        from .integrations import get_crm_client

        client = get_crm_client()
        results = client.search_customers(keyword)

        if not results:
            return f"未找到匹配 '{keyword}' 的客户"

        lines = [f"找到 {len(results)} 位匹配客户:\n"]
        for c in results:
            lines.append(f"  - {c.customer_id} | {c.name} | {c.company} | {c.level} | {c.status}")
        return "\n".join(lines)

    except Exception as e:
        logger.exception("CRM 客户搜索失败")
        return f"CRM 客户搜索失败: {sanitize_exception(e)}"


# 7. 业务系统集成工具：CRM 订单查询
@tool
@log_execution
def crm_query_order(order_id: str):
    """
    从 CRM 系统查询订单详情。
    当用户询问订单状态、订单详情时使用此工具。

    Args:
        order_id: 订单号（如 ORD-2024-001）
    """
    try:
        from .integrations import get_crm_client

        client = get_crm_client()
        order = client.get_order(order_id)

        if order is None:
            return f"❌ 未找到订单: {order_id}"

        return (
            f"📦 订单详情\n"
            f"  订单号: {order.order_id}\n"
            f"  客户ID: {order.customer_id}\n"
            f"  产品: {order.product}\n"
            f"  金额: ¥{order.amount:.2f}\n"
            f"  状态: {order.status}\n"
            f"  下单日期: {order.created_at}"
        )

    except Exception as e:
        logger.exception("CRM 订单查询失败")
        return f"CRM 订单查询失败: {sanitize_exception(e)}"


# 8. 业务系统集成工具：企业微信通知
@tool
@log_execution
def send_wecom_notification(message: str, msg_type: str = "text"):
    """
    通过企业微信群机器人发送通知消息。
    当用户要求发送企业微信通知、群消息推送时使用此工具。

    Args:
        message: 消息内容
        msg_type: 消息类型（text 或 markdown），默认 text
    """
    try:
        from .integrations import get_notifier

        notifier = get_notifier()

        if msg_type.lower() == "markdown":
            result = notifier.send_markdown(message)
        else:
            result = notifier.send_text(message)

        if result.get("errcode") == 0:
            return "✅ 企业微信通知发送成功"
        elif result.get("errcode") == -1:
            return "⚠️ 企业微信 Webhook 未配置，消息未发送。请在 .env 中设置 WECOM_WEBHOOK_URL"
        else:
            return f"❌ 发送失败: {result.get('errmsg', '未知错误')}"

    except Exception as e:
        logger.exception("企业微信通知发送失败")
        return f"企业微信通知发送失败: {sanitize_exception(e)}"


# 9. 业务系统集成工具：邮件发送
@tool
@log_execution
def send_email(to: str, subject: str, body: str, html: bool = False):
    """
    通过 SMTP 发送邮件。
    当用户要求发送邮件、邮件通知时使用此工具。

    Args:
        to: 收件人邮箱地址（多个地址用逗号分隔）
        subject: 邮件主题
        body: 邮件内容
        html: 是否为 HTML 格式（默认为纯文本）
    """
    try:
        from .integrations import get_email_sender

        sender = get_email_sender()

        to_addrs = [addr.strip() for addr in to.split(",") if addr.strip()]
        if not to_addrs:
            return "❌ 收件人地址为空"

        result = sender.send(to_addrs=to_addrs, subject=subject, body=body, html=html)

        if result.get("success"):
            return f"✅ 邮件发送成功: 收件人 {result['to']}, 主题: {result['subject']}"
        else:
            return f"❌ 邮件发送失败: {result.get('error', '未知错误')}"

    except Exception as e:
        logger.exception("邮件发送失败")
        return f"邮件发送失败: {sanitize_exception(e)}"


# ============================================================
# 11. P3 新增工具：calculate / fetch_url / read_file / write_file / python_repl
# ============================================================


# ---- 11.1 calculate：安全数学表达式求值 ----

import ast as _ast
import operator as _operator

# 允许的 AST 节点类型 + 运算符映射（白名单机制，杜绝任意代码执行）
# 不允许：Call（函数调用）、Attribute（属性访问）、Subscript（下标）、Import 等
# 注：ast.Num 已弃用（Python 3.8+ 用 ast.Constant），仅保留 Constant
_CALC_ALLOWED_NODES = {
    _ast.Expression,
    _ast.BinOp,
    _ast.UnaryOp,
    _ast.Constant,
    _ast.Add,
    _ast.Sub,
    _ast.Mult,
    _ast.Div,
    _ast.FloorDiv,
    _ast.Mod,
    _ast.Pow,
    _ast.USub,
    _ast.UAdd,
}
_CALC_OPERATORS = {
    _ast.Add: _operator.add,
    _ast.Sub: _operator.sub,
    _ast.Mult: _operator.mul,
    _ast.Div: _operator.truediv,
    _ast.FloorDiv: _operator.floordiv,
    _ast.Mod: _operator.mod,
    _ast.Pow: _operator.pow,
    _ast.USub: _operator.neg,
    _ast.UAdd: _operator.pos,
}

# 数学常量白名单（不允许访问 math 模块，仅暴露常用常量）
import math as _math

_CALC_CONSTANTS = {
    "pi": _math.pi,
    "e": _math.e,
    "tau": _math.tau,
}


def _safe_eval_node(node):
    """递归求值 AST 节点（白名单机制）"""
    if isinstance(node, _ast.Expression):
        return _safe_eval_node(node.body)
    if isinstance(node, _ast.Constant):  # Python 3.8+
        if isinstance(node.value, (int, float)):
            return node.value
        raise ValueError(f"不支持的常量类型: {type(node.value).__name__}")
    if isinstance(node, _ast.BinOp):
        op_type = type(node.op)
        if op_type not in _CALC_OPERATORS:
            raise ValueError(f"不支持的二元运算: {op_type.__name__}")
        left = _safe_eval_node(node.left)
        right = _safe_eval_node(node.right)
        return _CALC_OPERATORS[op_type](left, right)
    if isinstance(node, _ast.UnaryOp):
        op_type = type(node.op)
        if op_type not in _CALC_OPERATORS:
            raise ValueError(f"不支持的一元运算: {op_type.__name__}")
        operand = _safe_eval_node(node.operand)
        return _CALC_OPERATORS[op_type](operand)
    # 显式拒绝危险节点
    if isinstance(node, (_ast.Call, _ast.Attribute, _ast.Subscript, _ast.Name)):
        raise ValueError(f"禁止访问: {type(node).__name__}")
    raise ValueError(f"不支持的语法节点: {type(node).__name__}")


@tool
def calculate(expression: str) -> str:
    """安全数学表达式求值，支持 + - * / // % ** 和一元正负号，以及常量 pi/e/tau。

    用于精确计算（避免 LLM 心算出错）。不支持函数调用、变量、属性访问。

    Args:
        expression: 数学表达式，如 "2 + 3 * 4"、"3.14 * 2 ** 10"、"pi * 2"

    Returns:
        求值结果字符串；表达式非法时返回错误提示
    """
    if not expression or not expression.strip():
        return "错误：表达式为空"
    expr = expression.strip()
    # 长度限制，防止构造超长表达式 DoS
    if len(expr) > 200:
        return "错误：表达式过长（>200 字符）"

    try:
        # 解析为 AST
        tree = _ast.parse(expr, mode="eval")
        # 替换 Name 节点（pi/e/tau）为常量值
        for node in _ast.walk(tree):
            if isinstance(node, _ast.Name):
                if node.id in _CALC_CONSTANTS:
                    # 用 Constant 替换 Name（Python 3.8+）
                    node.__class__ = _ast.Constant
                    node.value = _CALC_CONSTANTS[node.id]
                    if hasattr(node, "id"):
                        del node.id
                    if hasattr(node, "ctx"):
                        del node.ctx
                else:
                    return f"错误：未知变量 '{node.id}'（仅支持 pi/e/tau）"
        result = _safe_eval_node(tree)
        # 整数结果去掉小数点
        if isinstance(result, float) and result.is_integer():
            result = int(result)
        return str(result)
    except (ValueError, TypeError, ZeroDivisionError) as e:
        return f"计算错误: {e}"
    except SyntaxError:
        return "错误：表达式语法无效"
    except Exception as e:
        logger.exception("calculate 工具异常")
        return f"计算异常: {sanitize_exception(e)}"


# ---- 11.2 fetch_url：网页内容抓取 ----


def _check_ssrf(url: str) -> str | None:
    """P0 SSRF 防护：校验 URL 目标 IP 是否为内网/回环/链路本地/保留地址。

    Args:
        url: 待校验的 URL

    Returns:
        错误消息字符串（拒绝访问）；None 表示通过校验
    """
    import ipaddress
    import socket
    from urllib.parse import urlparse

    parsed = urlparse(url)
    if not parsed.hostname:
        return "错误：URL 缺少主机名"

    try:
        # 解析主机名为 IP 地址（可能返回多个 A/AAAA 记录）
        addrinfos = socket.getaddrinfo(parsed.hostname, None)
    except socket.gaierror:
        # DNS 解析失败：交由后续 httpx 请求处理（会返回请求错误）
        return None

    for _, _, _, _, sockaddr in addrinfos:
        ip = ipaddress.ip_address(sockaddr[0])
        # 拒绝所有非公网地址：回环/私有/链路本地/保留/多播/未分配
        if (
            ip.is_loopback
            or ip.is_private
            or ip.is_link_local
            or ip.is_reserved
            or ip.is_multicast
            or ip.is_unspecified
        ):
            return (
                f"错误：禁止访问非公网地址 {ip}（SSRF 防护）。"
                "fetch_url 仅允许访问公网 URL。"
            )
    return None


@tool
def fetch_url(url: str) -> str:
    """抓取指定 URL 的网页正文内容，自动去除 HTML 标签和导航栏等噪音。

    配合 web_search 使用：web_search 找到相关链接，fetch_url 读取完整内容。
    仅支持 HTTP/HTTPS，限制响应体大小防止内存耗尽。

    Args:
        url: 目标 URL（必须以 http:// 或 https:// 开头）

    Returns:
        网页正文文本（截断到合理长度）；失败时返回错误提示
    """
    if not url or not url.strip():
        return "错误：URL 为空"
    url = url.strip()
    if not (url.startswith("http://") or url.startswith("https://")):
        return "错误：URL 必须以 http:// 或 https:// 开头"
    if len(url) > 2048:
        return "错误：URL 过长（>2048 字符）"

    # P0 安全修复：SSRF 防护 - 拒绝内网/回环/链路本地/元数据地址
    # 防止攻击者通过 prompt injection 让 agent 抓取：
    #   - http://169.254.169.254/... (云厂商元数据服务，可窃取 IAM 临时凭证)
    #   - http://127.0.0.1:8000/health/deep (内网管理面板)
    #   - http://10.0.0.1/... / http://192.168.0.1/... (内网横向渗透)
    err = _check_ssrf(url)
    if err:
        return err

    settings = get_settings()
    import httpx

    try:
        with httpx.Client(
            timeout=float(settings.tools_fetch_url_timeout),
            # P0 SSRF 防护：禁用自动重定向，防止外部服务器 302 跳转到内网地址绕过初始校验
            # 由我们手动校验每一跳的目标
            follow_redirects=False,
        ) as client:
            # P0 SSRF：手动跟随重定向，每一跳都校验目标 URL
            current_url = url
            for _redir in range(5):  # 最多 5 次重定向
                resp = client.get(
                    current_url,
                    headers={"User-Agent": settings.tools_fetch_url_user_agent},
                )
                if resp.is_redirect:
                    location = resp.headers.get("location", "")
                    if not location:
                        break
                    # 拼接相对重定向 URL
                    import urllib.parse as _up

                    next_url = _up.urljoin(current_url, location)
                    # 每跳都校验目标 IP，防止 302 跳到内网
                    redir_err = _check_ssrf(next_url)
                    if redir_err:
                        return redir_err
                    current_url = next_url
                    continue
                break
            resp.raise_for_status()

            # 检查 Content-Length，超过限制直接拒绝
            content_length = resp.headers.get("Content-Length")
            if content_length and int(content_length) > settings.tools_fetch_url_max_size:
                return f"错误：响应体过大（>{settings.tools_fetch_url_max_size} 字节）"

            content = resp.text
            if len(content) > settings.tools_fetch_url_max_size:
                content = content[: settings.tools_fetch_url_max_size]

        # 提取正文：优先用 readability-lxml，回退到粗暴去标签
        try:
            from readability import Document

            doc = Document(content)
            title = doc.short_title() or ""
            text = doc.summary()
            # 去 HTML 标签
            import re

            text = re.sub(r"<[^>]+>", " ", text)
            text = re.sub(r"\s+", " ", text).strip()
            result = f"标题: {title}\n\n{text}" if title else text
        except ImportError:
            # readability 未安装，回退到粗暴去标签
            import re

            text = re.sub(r"<script[^>]*>.*?</script>", "", content, flags=re.DOTALL | re.IGNORECASE)
            text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL | re.IGNORECASE)
            text = re.sub(r"<[^>]+>", " ", text)
            text = re.sub(r"\s+", " ", text).strip()
            result = text

        # 截断到合理长度，避免打满 LLM 上下文
        max_chars = 8000
        if len(result) > max_chars:
            result = result[:max_chars] + "\n\n[内容已截断]"
        return result
    except httpx.HTTPStatusError as e:
        return f"HTTP 错误: {e.response.status_code} {e.response.reason_phrase}"
    except httpx.RequestError as e:
        return f"请求失败: {sanitize_exception(e)}"
    except Exception as e:
        logger.exception("fetch_url 工具异常")
        return f"抓取异常: {sanitize_exception(e)}"


# ---- 11.3 read_file / write_file：文件读写（路径白名单）----


def _validate_workspace_path(file_path: str, must_exist: bool = False):
    """校验文件路径是否在 workspace 目录内，返回绝对 Path 对象

    防越权核心：通过 resolve() 解析符号链接后，检查是否在 workspace 内
    防止 ../、符号链接等绕过

    Args:
        file_path: 用户提供的相对或绝对路径
        must_exist: True 时要求文件必须存在

    Raises:
        ValueError: workspace 未配置 / 路径越界 / 文件不存在
    """
    from pathlib import Path

    settings = get_settings()
    workspace = settings.tools_workspace_dir
    if not workspace:
        raise ValueError("文件工具未启用（TOOLS_WORKSPACE_DIR 未配置）")

    workspace_root = Path(workspace).resolve()
    if not workspace_root.exists():
        raise ValueError(f"workspace 目录不存在: {workspace}")

    # 拼接并解析为绝对路径（resolve 会展开符号链接）
    target = (workspace_root / file_path).resolve()

    # P0 安全修复：用 relative_to 替代 str.startswith，消除前缀混淆漏洞
    # 旧实现 `str(target).startswith(str(workspace_root))` 在 workspace=/data/workspace
    # 而 target=/data/workspace-evil/secret.txt 时返回 True，导致路径穿越可读写任意文件
    # relative_to 在 target 不在 workspace 子树内时抛 ValueError，天然安全
    try:
        target.relative_to(workspace_root)
    except ValueError:
        raise ValueError(f"路径越界：{file_path} 不在 workspace 内")

    # 防止读取目录
    if must_exist and not target.is_file():
        raise ValueError(f"文件不存在或不是普通文件: {file_path}")

    return target


@tool
def read_file(file_path: str) -> str:
    """读取 workspace 目录内的文本文件内容。

    仅能读取 TOOLS_WORKSPACE_DIR 配置目录内的文件，防止越权访问系统文件。
    自动检测 UTF-8/GBK 编码；二进制文件返回提示而非乱码。

    Args:
        file_path: 相对于 workspace 的文件路径，如 "data/test.txt"

    Returns:
        文件文本内容（截断到合理长度）；失败时返回错误提示
    """
    try:
        target = _validate_workspace_path(file_path, must_exist=True)
        # 限制单文件读取大小（防止读 GB 级日志打满上下文）
        max_size = 512 * 1024  # 512KB
        size = target.stat().st_size
        if size > max_size:
            return f"错误：文件过大（{size} 字节，> {max_size}）"

        # 尝试 UTF-8，回退 GBK（Windows 中文环境兼容）
        for encoding in ("utf-8", "gbk", "latin-1"):
            try:
                content = target.read_text(encoding=encoding)
                # 截断到合理长度
                if len(content) > 32000:
                    content = content[:32000] + "\n\n[文件已截断]"
                return content
            except UnicodeDecodeError:
                continue
        return "错误：无法解码文件（非文本文件？）"
    except ValueError as e:
        return f"错误: {e}"
    except OSError as e:
        return f"读取失败: {sanitize_exception(e)}"
    except Exception as e:
        logger.exception("read_file 工具异常")
        return f"读取异常: {sanitize_exception(e)}"


@tool
def write_file(file_path: str, content: str) -> str:
    """向 workspace 目录写入文本文件（覆盖已存在的文件）。

    仅能写入 TOOLS_WORKSPACE_DIR 配置目录内，防止越权写入系统文件。
    自动创建父目录。

    Args:
        file_path: 相对于 workspace 的文件路径
        content: 要写入的文本内容

    Returns:
        成功时返回写入的字符数；失败时返回错误提示
    """
    if content is None:
        content = ""
    # 限制写入大小，防止 LLM 生成超大文件耗尽磁盘
    if len(content) > 1024 * 1024:  # 1MB
        return f"错误：内容过大（{len(content)} 字符，> 1MB）"

    try:
        target = _validate_workspace_path(file_path, must_exist=False)
        # 自动创建父目录
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(content, encoding="utf-8")
        return f"已写入 {len(content)} 字符到 {file_path}"
    except ValueError as e:
        return f"错误: {e}"
    except OSError as e:
        return f"写入失败: {sanitize_exception(e)}"
    except Exception as e:
        logger.exception("write_file 工具异常")
        return f"写入异常: {sanitize_exception(e)}"


# ---- 11.4 python_repl：受控 Python 执行 ----


@tool
def python_repl(code: str) -> str:
    """在受限沙箱中执行 Python 代码，返回 stdout 输出。

    用于数据分析、字符串处理、复杂计算等场景。
    已内置 math、statistics、json、re、datetime、itertools 模块。
    禁止：import、open、exec、eval、__import__、subprocess 等。
    执行超时默认 10 秒，输出截断到 4096 字符。

    Args:
        code: Python 代码（支持多行）

    Returns:
        stdout 输出；执行错误时返回异常信息
    """
    if not code or not code.strip():
        return "错误：代码为空"
    if len(code) > 10000:
        return "错误：代码过长（>10000 字符）"

    settings = get_settings()
    timeout = float(settings.tools_python_repl_timeout)
    max_output = settings.tools_python_repl_max_output

    # 危险关键字黑名单（pre-check，主防护靠沙箱内置 __builtins__）
    # P0 安全加固：补齐 Python 沙箱逃逸常用的 dunder 属性，防止 __subclasses__ 绕过
    # 旧黑名单遗漏 __class__ / __subclasses__ / __base__ / __mro__ / __globals__ 等，
    # 攻击者可通过 ().__class__.__base__.__subclasses__() 枚举所有已加载类
    # （含 subprocess.Popen、os._wrap_close 等）实现 RCE
    _DANGEROUS = (
        "import ",
        "import(",
        "__import__",
        "__builtins__",
        "subprocess",
        "os.system",
        "os.popen",
        "os.exec",
        "open(",
        "exec(",
        "eval(",
        "compile(",
        "globals(",
        "locals(",
        "vars(",
        # P0：沙箱逃逸常用 dunder 属性（覆盖 __subclasses__ 链 + __globals__ 链）
        "__class__",
        "__subclasses__",
        "__base__",
        "__bases__",
        "__mro__",
        "__globals__",
        "__code__",
        "__reduce__",
        "__reduce_ex__",
        "__getstate__",
        "__setstate__",
        "__dict__",
        "__init_subclass__",
        "__subclasshook__",
    )
    code_lower = code.lower()
    for kw in _DANGEROUS:
        if kw in code_lower:
            return f"错误：禁止使用危险关键字 '{kw.strip()}'"

    import io
    import signal
    from contextlib import redirect_stderr, redirect_stdout

    # 受限内建：仅暴露安全函数
    _SAFE_BUILTINS = {
        "print": print,
        "len": len,
        "range": range,
        "enumerate": enumerate,
        "zip": zip,
        "map": map,
        "filter": filter,
        "sum": sum,
        "min": min,
        "max": max,
        "abs": abs,
        "round": round,
        "sorted": sorted,
        "reversed": reversed,
        "list": list,
        "dict": dict,
        "set": set,
        "tuple": tuple,
        "str": str,
        "int": int,
        "float": float,
        "bool": bool,
        "type": type,
        "isinstance": isinstance,
        "all": all,
        "any": any,
        "True": True,
        "False": False,
        "None": None,
    }
    # 允许的模块（预先 import，注入到 globals）
    import collections
    import datetime
    import itertools
    import json
    import math
    import re
    import statistics

    _SAFE_GLOBALS = {
        "__builtins__": _SAFE_BUILTINS,
        "math": math,
        "statistics": statistics,
        "json": json,
        "re": re,
        "datetime": datetime,
        "itertools": itertools,
        "collections": collections,
    }

    stdout_buf = io.StringIO()
    stderr_buf = io.StringIO()

    def _run():
        with redirect_stdout(stdout_buf), redirect_stderr(stderr_buf):
            exec(code, _SAFE_GLOBALS, {})

    # P0 安全加固：执行前 AST 静态分析，拒绝明显的死循环模式（DoS 防护）
    # 检测模式：while <truthy constant> 且循环体内无 break 语句
    # 这是跨平台兜底（Unix 有 SIGALRM 可中断，Windows 无信号机制需静态防护）
    try:
        tree = _ast.parse(code)
        for node in _ast.walk(tree):
            if isinstance(node, _ast.While):
                # 判断条件是否为常量真值（True / 非零数字 / 非空字符串）
                cond = node.test
                is_truthy_const = False
                if isinstance(cond, _ast.Constant):
                    if isinstance(cond.value, (int, float, bool)):
                        is_truthy_const = bool(cond.value)
                    elif isinstance(cond.value, str):
                        is_truthy_const = bool(cond.value)
                if is_truthy_const:
                    # 检查循环体内是否有 break（有 break 则允许）
                    has_break = any(isinstance(child, _ast.Break) for child in _ast.walk(node))
                    if not has_break:
                        return (
                            "错误: 检测到无限循环（while <truthy> 无 break），"
                            "为防止资源耗尽已拒绝执行。请改用带退出条件的循环。"
                        )
            # P0 安全加固：禁止访问以下划线开头的属性（拦截 __subclasses__ 链 + 反射绕过）
            # 这是关键字黑名单的 AST 级补充：黑名单可被字符串拼接绕过
            # （如 getattr(x, "__cla"+"ss__")），AST 检测覆盖所有 Attribute 节点
            if isinstance(node, _ast.Attribute) and node.attr.startswith("_"):
                return (
                    "错误: 禁止访问以下划线开头的属性（防止沙箱逃逸）。"
                    "如需数据处理请使用提供的 math/statistics/json 等模块。"
                )
            # P0：拦截反射函数调用（getattr/setattr/hasattr/delattr 可绕过 Attribute 检测）
            if isinstance(node, _ast.Call) and isinstance(node.func, _ast.Name):
                if node.func.id in ("getattr", "setattr", "hasattr", "delattr", "vars", "dir"):
                    return (
                        f"错误: 禁止使用反射函数 {node.func.id}（防止沙箱逃逸）。"
                    )
    except SyntaxError:
        # 语法错误交给后续 exec 抛出，不在此处拦截
        pass

    # 信号超时（仅 Unix 有效，Windows 用线程兜底）
    try:
        if hasattr(signal, "SIGALRM"):

            def _handler(signum, frame):
                raise TimeoutError(f"执行超时（{timeout}s）")

            old = signal.signal(signal.SIGALRM, _handler)
            signal.setitimer(signal.ITIMER_REAL, timeout)
            try:
                _run()
            finally:
                signal.setitimer(signal.ITIMER_REAL, 0)
                signal.signal(signal.SIGALRM, old)
        else:
            # Windows：用 threading.Timer 兜底（无法强杀线程，仅作提示）
            # P0 加固：结合上方的 AST 静态分析预检，可拦截绝大多数死循环 DoS 场景
            # 残余风险（如递归无限调用）由 tools_python_repl_timeout 配置的 Timer 提示
            import threading

            timed_out = {"flag": False}

            def _watchdog():
                timed_out["flag"] = True

            timer = threading.Timer(timeout, _watchdog)
            timer.start()
            try:
                _run()
            finally:
                timer.cancel()
            if timed_out["flag"]:
                raise TimeoutError(f"执行超时（{timeout}s）")
    except TimeoutError as e:
        return f"错误: {e}"
    except Exception as e:
        err = stderr_buf.getvalue()
        return f"执行错误: {sanitize_exception(e)}" + (f"\n{err}" if err else "")

    output = stdout_buf.getvalue()
    if not output:
        output = "(无输出)"
    if len(output) > max_output:
        output = output[:max_output] + f"\n\n[输出已截断，共 {len(output)} 字符]"
    return output


# ---- 11.5 parse_pdf / parse_excel / parse_csv：文件解析工具 ----


@tool
def parse_pdf(file_path: str) -> str:
    """解析 PDF 文件，提取全部文本内容。

    仅能解析 TOOLS_WORKSPACE_DIR 目录内的文件，防止越权访问系统文件。
    支持多页 PDF，自动合并各页文本；无法解析的扫描件（图片型 PDF）返回提示。

    Args:
        file_path: 相对于 workspace 的 PDF 文件路径

    Returns:
        PDF 文本内容（截断到合理长度）；失败时返回错误提示
    """
    try:
        target = _validate_workspace_path(file_path, must_exist=True)
        # 限制文件大小（PDF 解析内存占用较高）
        max_size = 10 * 1024 * 1024  # 10MB
        if target.stat().st_size > max_size:
            return f"错误：PDF 过大（>{max_size} 字节）"

        from pypdf import PdfReader

        reader = PdfReader(str(target))
        pages = []
        for i, page in enumerate(reader.pages):
            try:
                text = page.extract_text() or ""
                if text.strip():
                    pages.append(f"--- 第 {i + 1} 页 ---\n{text.strip()}")
            except Exception:
                logger.warning("PDF 第 %d 页解析失败", i + 1, exc_info=True)
                continue

        if not pages:
            return "提示：未提取到文本（可能是扫描件/图片型 PDF，需 OCR 支持）"

        result = "\n\n".join(pages)
        # 截断到合理长度
        if len(result) > 32000:
            result = result[:32000] + "\n\n[内容已截断]"
        return result
    except ValueError as e:
        return f"错误: {e}"
    except Exception as e:
        logger.exception("parse_pdf 工具异常")
        return f"PDF 解析异常: {sanitize_exception(e)}"


@tool
def parse_excel(file_path: str, sheet_name: str = "") -> str:
    """解析 Excel 文件（.xlsx/.xls），返回指定工作表或全部数据。

    仅能解析 TOOLS_WORKSPACE_DIR 目录内的文件。
    输出为 Markdown 表格格式，便于 LLM 理解；自动跳过空行。

    Args:
        file_path: 相对于 workspace 的 Excel 文件路径
        sheet_name: 指定工作表名，为空则解析第一个工作表

    Returns:
        Markdown 表格格式的数据；失败时返回错误提示
    """
    try:
        target = _validate_workspace_path(file_path, must_exist=True)
        max_size = 10 * 1024 * 1024  # 10MB
        if target.stat().st_size > max_size:
            return f"错误：Excel 过大（>{max_size} 字节）"

        from openpyxl import load_workbook

        wb = load_workbook(str(target), read_only=True, data_only=True)
        try:
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return f"错误：工作表 '{sheet_name}' 不存在，可用: {wb.sheetnames}"
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return f"工作表 '{sheet_name}' 为空"

            # 转换为 Markdown 表格
            lines = [f"### 工作表: {sheet_name}\n"]
            # 表头
            header = [str(c) if c is not None else "" for c in rows[0]]
            lines.append("| " + " | ".join(header) + " |")
            lines.append("| " + " | ".join(["---"] * len(header)) + " |")
            # 数据行（最多 100 行，防止超大表打满上下文）
            for row in rows[1:101]:
                cells = [str(c) if c is not None else "" for c in row]
                lines.append("| " + " | ".join(cells) + " |")

            if len(rows) > 101:
                lines.append(f"\n(共 {len(rows) - 1} 行数据，已截断显示前 100 行)")

            result = "\n".join(lines)
            if len(result) > 32000:
                result = result[:32000] + "\n\n[内容已截断]"
            return result
        finally:
            wb.close()
    except ValueError as e:
        return f"错误: {e}"
    except ImportError:
        return "错误：openpyxl 未安装，无法解析 Excel"
    except Exception as e:
        logger.exception("parse_excel 工具异常")
        return f"Excel 解析异常: {sanitize_exception(e)}"


@tool
def parse_csv(file_path: str, delimiter: str = ",") -> str:
    """解析 CSV 文件，返回 Markdown 表格格式的数据。

    仅能解析 TOOLS_WORKSPACE_DIR 目录内的文件。
    自动检测表头，跳过空行；支持自定义分隔符（如 TSV 用 \\t）。

    Args:
        file_path: 相对于 workspace 的 CSV 文件路径
        delimiter: 字段分隔符，默认逗号

    Returns:
        Markdown 表格格式的数据；失败时返回错误提示
    """
    try:
        target = _validate_workspace_path(file_path, must_exist=True)
        max_size = 10 * 1024 * 1024  # 10MB
        if target.stat().st_size > max_size:
            return f"错误：CSV 过大（>{max_size} 字节）"

        import csv

        # 尝试 UTF-8，回退 GBK（Windows 中文环境）
        content = None
        for encoding in ("utf-8-sig", "utf-8", "gbk", "latin-1"):
            try:
                content = target.read_text(encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        if content is None:
            return "错误：无法解码文件"

        # 处理 \t 转义
        if delimiter == "\\t":
            delimiter = "\t"

        reader = csv.reader(content.splitlines(), delimiter=delimiter)
        rows = list(reader)
        if not rows:
            return "文件为空"

        # 过滤空行
        rows = [r for r in rows if any(c.strip() for c in r)]
        if not rows:
            return "文件无有效数据"

        # 转换为 Markdown 表格
        lines = []
        header = [c.strip() for c in rows[0]]
        lines.append("| " + " | ".join(header) + " |")
        lines.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in rows[1:101]:
            cells = [c.strip() for c in row]
            # 补齐列数
            while len(cells) < len(header):
                cells.append("")
            lines.append("| " + " | ".join(cells) + " |")

        if len(rows) > 101:
            lines.append(f"\n(共 {len(rows) - 1} 行数据，已截断显示前 100 行)")

        result = "\n".join(lines)
        if len(result) > 32000:
            result = result[:32000] + "\n\n[内容已截断]"
        return result
    except ValueError as e:
        return f"错误: {e}"
    except Exception as e:
        logger.exception("parse_csv 工具异常")
        return f"CSV 解析异常: {sanitize_exception(e)}"


# ---- 11.6 generate_qrcode：二维码生成（离线）----


@tool
def generate_qrcode(data: str, file_path: str = "qrcode.png") -> str:
    """生成二维码图片并保存到 workspace 目录。

    离线实现，无需外部 API。适合生成 URL/文本/名片等二维码。

    Args:
        data: 要编码的数据（URL/文本等）
        file_path: 保存路径（相对于 workspace），默认 qrcode.png

    Returns:
        成功时返回保存路径与数据长度；失败时返回错误提示
    """
    if not data or not data.strip():
        return "错误：数据为空"
    if len(data) > 2000:
        return f"错误：数据过长（{len(data)} 字符，>2000，二维码可能无法识别）"

    try:
        target = _validate_workspace_path(file_path, must_exist=False)
        target.parent.mkdir(parents=True, exist_ok=True)

        try:
            import qrcode
        except ImportError:
            return "错误：qrcode 库未安装"

        # 容错级别 M（约 15% 数据冗余），适合大多数场景
        qr = qrcode.QRCode(
            version=None,  # 自动选择版本
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=4,
        )
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        img.save(str(target))
        return f"已生成二维码到 {file_path}（数据长度 {len(data)} 字符）"
    except ValueError as e:
        return f"错误: {e}"
    except Exception as e:
        logger.exception("generate_qrcode 工具异常")
        return f"二维码生成异常: {sanitize_exception(e)}"


# ---- 11.7 get_weather / get_exchange_rate / translate：外部信息查询 ----


@tool
def get_weather(location: str) -> str:
    """查询指定城市的实时天气（和风天气 API）。

    需配置 WEATHER_API_KEY 与 WEATHER_API_BASE（默认和风天气）。
    支持城市名（如 "北京"）或 location ID。

    Args:
        location: 城市名（如 "北京"、"上海"）或和风天气 location ID

    Returns:
        天气信息（温度/天气现象/风力/湿度等）；失败时返回错误提示
    """
    if not location or not location.strip():
        return "错误：城市为空"
    location = location.strip()
    if len(location) > 100:
        return "错误：城市名过长"

    settings = get_settings()
    api_key = settings.weather_api_key
    if not api_key:
        return "错误：天气服务未配置（WEATHER_API_KEY 未设置）"

    base_url = settings.weather_api_base or "https://devapi.qweather.com/v7"
    import httpx

    try:
        with httpx.Client(timeout=10.0) as client:
            # 步骤 1：城市名 → location ID（和风天气 Geo API）
            if not location.isdigit():
                geo_resp = client.get(
                    "https://geoapi.qweather.com/v2/city/lookup",
                    params={"location": location, "key": api_key},
                )
                geo_resp.raise_for_status()
                geo_data = geo_resp.json()
                if geo_data.get("code") != "200" or not geo_data.get("location"):
                    return f"错误：未找到城市 '{location}'"
                location_id = geo_data["location"][0]["id"]
                city_name = geo_data["location"][0]["name"]
            else:
                location_id = location
                city_name = location

            # 步骤 2：查询实时天气
            weather_resp = client.get(
                f"{base_url}/weather/now",
                params={"location": location_id, "key": api_key},
            )
            weather_resp.raise_for_status()
            weather_data = weather_resp.json()
            if weather_data.get("code") != "200":
                return f"错误：天气查询失败（code={weather_data.get('code')}）"

            now = weather_data["now"]
            result = (
                f"城市: {city_name}\n"
                f"温度: {now['temp']}°C（体感 {now['feelsLike']}°C）\n"
                f"天气: {now['text']}\n"
                f"风向风力: {now['windDir']} {now['windScale']}级\n"
                f"湿度: {now['humidity']}%\n"
                f"观测时间: {now['obsTime']}"
            )
            return result
    except httpx.HTTPStatusError as e:
        return f"HTTP 错误: {e.response.status_code}"
    except httpx.RequestError as e:
        return f"请求失败: {sanitize_exception(e)}"
    except Exception as e:
        logger.exception("get_weather 工具异常")
        return f"天气查询异常: {sanitize_exception(e)}"


@tool
def get_exchange_rate(base: str = "USD", target: str = "CNY") -> str:
    """查询实时汇率（fixer.io API 或央行汇率）。

    需配置 EXCHANGE_RATE_API_KEY。默认使用 fixer.io（欧元基准），
    自动换算为指定基准货币。

    Args:
        base: 基准货币代码（如 USD/EUR/JPY），默认 USD
        target: 目标货币代码（如 CNY/USD/EUR），默认 CNY

    Returns:
        汇率信息（1 base = X target）；失败时返回错误提示
    """
    base = base.upper().strip()
    target = target.upper().strip()
    if not base or not target:
        return "错误：货币代码为空"
    if len(base) != 3 or len(target) != 3:
        return "错误：货币代码应为 3 字符（如 USD/CNY/EUR）"

    settings = get_settings()
    api_key = settings.exchange_rate_api_key
    if not api_key:
        return "错误：汇率服务未配置（EXCHANGE_RATE_API_KEY 未设置）"

    base_url = settings.exchange_rate_api_base or "https://data.fixer.io/api"
    import httpx

    try:
        with httpx.Client(timeout=10.0) as client:
            resp = client.get(
                f"{base_url}/latest",
                params={
                    "access_key": api_key,
                    "base": "EUR",  # fixer.io 免费版仅支持 EUR 基准
                    "symbols": f"{base},{target}",
                },
            )
            resp.raise_for_status()
            data = resp.json()
            if not data.get("success"):
                return f"错误：汇率查询失败（{data.get('error', {})}）"

            rates = data.get("rates", {})
            if base not in rates or target not in rates:
                return f"错误：不支持的货币代码（base={base}, target={target}）"

            # fixer.io 基准为 EUR，需换算：1 base = rates[target] / rates[base] target
            rate = rates[target] / rates[base]
            date = data.get("date", "未知")
            result = f"汇率（{date}）:\n" f"1 {base} = {rate:.4f} {target}\n" f"1 {target} = {1 / rate:.4f} {base}"
            return result
    except httpx.HTTPStatusError as e:
        return f"HTTP 错误: {e.response.status_code}"
    except httpx.RequestError as e:
        return f"请求失败: {sanitize_exception(e)}"
    except Exception as e:
        logger.exception("get_exchange_rate 工具异常")
        return f"汇率查询异常: {sanitize_exception(e)}"


@tool
def translate(text: str, from_lang: str = "auto", to_lang: str = "zh") -> str:
    """翻译文本（百度翻译 API）。

    需配置 BAIDU_TRANSLATE_APP_ID 与 BAIDU_TRANSLATE_API_KEY。
    支持自动检测源语言。

    常用语言代码：zh（中文）、en（英文）、jp（日文）、kor（韩文）、fra（法文）、de（德文）。

    Args:
        text: 要翻译的文本（≤6000 字符）
        from_lang: 源语言代码，默认 auto（自动检测）
        to_lang: 目标语言代码，默认 zh（中文）

    Returns:
        翻译结果；失败时返回错误提示
    """
    if not text or not text.strip():
        return "错误：文本为空"
    text = text.strip()
    if len(text) > 6000:
        return f"错误：文本过长（{len(text)} 字符，>6000）"

    settings = get_settings()
    app_id = settings.baidu_translate_app_id
    api_key = settings.baidu_translate_api_key
    if not app_id or not api_key:
        return "错误：翻译服务未配置（BAIDU_TRANSLATE_APP_ID/API_KEY 未设置）"

    import hashlib
    import random

    import httpx

    # 百度翻译 API 签名：md5(appid + q + salt + key)
    salt = str(random.randint(32768, 65536))
    sign_str = app_id + text + salt + api_key
    sign = hashlib.md5(sign_str.encode("utf-8")).hexdigest()

    try:
        with httpx.Client(timeout=10.0) as client:
            resp = client.get(
                "https://fanyi-api.baidu.com/api/trans/vip/translate",
                params={
                    "q": text,
                    "from": from_lang,
                    "to": to_lang,
                    "appid": app_id,
                    "salt": salt,
                    "sign": sign,
                },
            )
            resp.raise_for_status()
            data = resp.json()

            # 百度 API 错误返回 {"error_code": "xxx", "error_msg": "..."}
            if "error_code" in data:
                return f"错误：翻译失败（{data.get('error_msg')}）"

            results = data.get("trans_result", [])
            if not results:
                return "错误：翻译结果为空"

            # 合并多段翻译结果
            translated = "\n".join(item.get("dst", "") for item in results)
            detected_from = data.get("from", from_lang)
            return f"源语言: {detected_from}\n目标语言: {to_lang}\n翻译结果:\n{translated}"
    except httpx.HTTPStatusError as e:
        return f"HTTP 错误: {e.response.status_code}"
    except httpx.RequestError as e:
        return f"请求失败: {sanitize_exception(e)}"
    except Exception as e:
        logger.exception("translate 工具异常")
        return f"翻译异常: {sanitize_exception(e)}"


# ---- 11.8 hash_encode / text_diff / regex_test / unit_convert：离线基础工具 ----


@tool
def hash_encode(text: str, algorithm: str = "md5") -> str:
    """计算文本的哈希值或进行编码转换。

    支持的算法：
    - 哈希：md5 / sha1 / sha256 / sha512（返回十六进制摘要）
    - 编码：base64（编码）/ base64_decode（解码）/ url_encode / url_decode / hex_encode / hex_decode

    用于数据指纹、签名生成、URL 处理、编码数据解码等场景。

    Args:
        text: 输入文本
        algorithm: 算法名，默认 md5

    Returns:
        哈希摘要或编码结果字符串；失败时返回错误提示
    """
    if not text:
        return "错误：输入为空"
    if len(text) > 1024 * 1024:  # 1MB
        return f"错误：输入过长（{len(text)} 字符，>1MB）"

    algorithm = algorithm.lower().strip()

    try:
        # 哈希算法
        if algorithm in ("md5", "sha1", "sha256", "sha512"):
            import hashlib

            h = hashlib.new(algorithm)
            h.update(text.encode("utf-8"))
            return h.hexdigest()

        # Base64 编解码
        if algorithm == "base64":
            import base64

            return base64.b64encode(text.encode("utf-8")).decode("ascii")
        if algorithm == "base64_decode":
            import base64

            return base64.b64decode(text).decode("utf-8")

        # URL 编解码
        if algorithm == "url_encode":
            from urllib.parse import quote

            return quote(text, safe="")
        if algorithm == "url_decode":
            from urllib.parse import unquote

            return unquote(text)

        # Hex 编解码
        if algorithm == "hex_encode":
            return text.encode("utf-8").hex()
        if algorithm == "hex_decode":
            return bytes.fromhex(text).decode("utf-8")

        return f"错误：不支持的算法 '{algorithm}'（支持: md5/sha1/sha256/sha512/base64/base64_decode/url_encode/url_decode/hex_encode/hex_decode）"
    except UnicodeDecodeError as e:
        return f"错误：解码失败（{e}）"
    except ValueError as e:
        return f"错误：{e}"
    except Exception as e:
        logger.exception("hash_encode 工具异常")
        return f"哈希/编码异常: {sanitize_exception(e)}"


@tool
def text_diff(text1: str, text2: str, lines_per_context: int = 3) -> str:
    """对比两段文本的差异，返回 unified diff 格式结果。

    用于文档版本对比、配置变更检查、代码 review 等场景。
    按行对比，支持自定义上下文行数。

    Args:
        text1: 原始文本
        text2: 修改后文本
        lines_per_context: 上下文行数（默认 3），影响 diff 输出中未变更行的显示数量

    Returns:
        unified diff 格式的差异结果；无差异时返回提示
    """
    if text1 is None:
        text1 = ""
    if text2 is None:
        text2 = ""
    if len(text1) > 100 * 1024 or len(text2) > 100 * 1024:
        return "错误：文本过长（>100KB）"

    import difflib

    lines1 = text1.splitlines(keepends=True)
    lines2 = text2.splitlines(keepends=True)

    # 限制上下文行数（difflib 要求 >=1）
    n = max(1, min(int(lines_per_context), 10))

    diff = difflib.unified_diff(
        lines1,
        lines2,
        fromfile="原始",
        tofile="修改后",
        n=n,
        lineterm="",
    )
    result = "\n".join(diff)

    if not result:
        return "两段文本完全相同，无差异"

    # 截断到合理长度
    if len(result) > 32000:
        result = result[:32000] + "\n\n[差异已截断]"
    return result


@tool
def regex_test(pattern: str, text: str, flags: str = "") -> str:
    """测试正则表达式并返回匹配结果。

    用于验证正则表达式、提取匹配内容、调试正则逻辑。
    支持常见 flags 组合（如 "i" 忽略大小写、"m" 多行模式）。

    Args:
        pattern: 正则表达式（Python re 语法）
        text: 待匹配文本
        flags: flags 字符串，支持的字符：i（IGNORECASE）、m（MULTILINE）、s（DOTALL）、x（VERBOSE）

    Returns:
        匹配结果（含捕获组）；无匹配时返回提示；正则错误时返回错误信息
    """
    if not pattern:
        return "错误：正则表达式为空"
    if not text:
        return "错误：待匹配文本为空"
    if len(pattern) > 1000:
        return "错误：正则表达式过长（>1000 字符）"
    if len(text) > 100 * 1024:
        return "错误：文本过长（>100KB）"

    import re

    # 解析 flags
    flag_map = {
        "i": re.IGNORECASE,
        "m": re.MULTILINE,
        "s": re.DOTALL,
        "x": re.VERBOSE,
    }
    re_flags = 0
    for f in flags.lower():
        if f in flag_map:
            re_flags |= flag_map[f]

    try:
        # 编译正则（捕获语法错误）
        compiled = re.compile(pattern, re_flags)
    except re.error as e:
        return f"正则语法错误: {e}"

    # 查找所有匹配
    matches = list(compiled.finditer(text))
    if not matches:
        return "无匹配"

    lines = [f"匹配数: {len(matches)}\n"]
    # 最多显示前 20 个匹配，防止超大输出
    for i, m in enumerate(matches[:20]):
        lines.append(f"--- 匹配 {i + 1} ---")
        lines.append(f"位置: {m.start()}-{m.end()}")
        lines.append(f"完整匹配: {m.group(0)}")
        # 显示捕获组
        if m.groups():
            for j, g in enumerate(m.groups(), 1):
                lines.append(f"捕获组 {j}: {g if g is not None else '(空)'}")
        # 显示命名捕获组
        if m.groupdict():
            for name, g in m.groupdict().items():
                lines.append(f"命名组 '{name}': {g if g is not None else '(空)'}")

    if len(matches) > 20:
        lines.append(f"\n(共 {len(matches)} 个匹配，已显示前 20 个)")

    result = "\n".join(lines)
    if len(result) > 8000:
        result = result[:8000] + "\n\n[结果已截断]"
    return result


@tool
def unit_convert(value: float, from_unit: str, to_unit: str, category: str = "length") -> str:
    """单位换算工具，支持长度/重量/温度/时间/数据量。

    用于跨境业务、技术文档翻译、数据计算辅助等场景，避免 LLM 单位换算出错。

    Args:
        value: 待换算的数值
        from_unit: 源单位（如 m/km/kg/C/s/KB）
        to_unit: 目标单位
        category: 类别，可选: length（长度）/ weight（重量）/ temperature（温度）/ time（时间）/ data（数据量）

    Returns:
        换算结果字符串；不支持的单位或类别时返回错误提示
    """
    try:
        value = float(value)
    except (TypeError, ValueError):
        return f"错误：value 不是有效数值（{value}）"

    from_unit = from_unit.lower().strip()
    to_unit = to_unit.lower().strip()
    category = category.lower().strip()

    if not from_unit or not to_unit:
        return "错误：单位为空"

    # 各类别的换算表（以基准单位为 1.0，其他单位为相对基准的倍数）
    # 基准单位：length=m, weight=kg, temperature=C, time=s, data=B
    conversion_tables = {
        "length": {
            "mm": 0.001,
            "cm": 0.01,
            "m": 1.0,
            "km": 1000.0,
            "in": 0.0254,
            "ft": 0.3048,
            "yd": 0.9144,
            "mi": 1609.344,
            "尺": 0.3333,
            "寸": 0.03333,
            "里": 500.0,
        },
        "weight": {
            "mg": 0.000001,
            "g": 0.001,
            "kg": 1.0,
            "t": 1000.0,
            "oz": 0.0283495,
            "lb": 0.453592,
            "斤": 0.5,
            "两": 0.05,
        },
        "time": {
            "ms": 0.001,
            "s": 1.0,
            "min": 60.0,
            "h": 3600.0,
            "day": 86400.0,
            "week": 604800.0,
        },
        "data": {
            "b": 1.0,
            "byte": 1.0,
            "kb": 1024.0,
            "mb": 1024**2,
            "gb": 1024**3,
            "tb": 1024**4,
            "pb": 1024**5,
            "kib": 1024.0,
            "mib": 1024**2,
            "gib": 1024**3,
        },
    }

    # 温度特殊处理（非比例换算，需公式）
    if category == "temperature":
        temp_units = {"c", "f", "k"}
        if from_unit not in temp_units or to_unit not in temp_units:
            return "错误：温度单位仅支持 c/f/k（摄氏/华氏/开尔文）"
        # 先转摄氏度
        if from_unit == "c":
            celsius = value
        elif from_unit == "f":
            celsius = (value - 32) * 5 / 9
        else:  # k
            celsius = value - 273.15
        # 再从摄氏度转目标
        if to_unit == "c":
            result = celsius
        elif to_unit == "f":
            result = celsius * 9 / 5 + 32
        else:  # k
            result = celsius + 273.15
        return f"{value} {from_unit.upper()} = {round(result, 4)} {to_unit.upper()}"

    # 其他类别：比例换算
    if category not in conversion_tables:
        return f"错误：不支持的类别 '{category}'（支持: length/weight/temperature/time/data）"

    table = conversion_tables[category]
    if from_unit not in table:
        return f"错误：{category} 类别不支持的源单位 '{from_unit}'（可用: {list(table.keys())}）"
    if to_unit not in table:
        return f"错误：{category} 类别不支持的目标单位 '{to_unit}'（可用: {list(table.keys())}）"

    # 换算：先转基准单位，再转目标单位
    base_value = value * table[from_unit]
    result = base_value / table[to_unit]
    # 智能精度：大数保留 2 位，小数保留 6 位
    if abs(result) >= 1000:
        result_str = f"{result:.2f}"
    else:
        result_str = f"{result:.6f}".rstrip("0").rstrip(".")
    return f"{value} {from_unit} = {result_str} {to_unit}"


# ============================================================
# 12. 汇总所有工具
# ============================================================
AGENT_TOOLS = [
    get_current_time,
    web_search,
    knowledge_search,
    knowledge_upload,
    crm_query_customer,
    crm_search_customers,
    crm_query_order,
    send_wecom_notification,
    send_email,
    # P3 新增工具 - 第一梯队
    calculate,
    fetch_url,
    read_file,
    write_file,
    python_repl,
    # P3 新增工具 - 第二梯队
    parse_pdf,
    parse_excel,
    parse_csv,
    generate_qrcode,
    get_weather,
    get_exchange_rate,
    translate,
    # P3 新增工具 - 第三梯队
    hash_encode,
    text_diff,
    regex_test,
    unit_convert,
]

# P0 工具权限分级：按 scope 暴露不同工具集，防止 readonly 用户通过 LLM 工具调用绕过 HTTP 层权限
# - readonly：仅只读工具（时间/搜索/知识库检索/CRM 查询/计算/URL抓取/文件读取/Python执行/文件解析/天气/汇率/翻译/哈希编码/文本对比/正则测试/单位换算）
# - write：readonly + knowledge_upload + write_file + generate_qrcode（写知识库/写文件/生成二维码）
# - admin：全部工具（含外发邮件/企业微信通知）
_READONLY_TOOLS = [
    get_current_time,
    web_search,
    knowledge_search,
    crm_query_customer,
    crm_search_customers,
    crm_query_order,
    # P3 第一梯队只读工具
    calculate,
    fetch_url,
    read_file,
    python_repl,
    # P3 第二梯队只读工具（解析/查询类）
    parse_pdf,
    parse_excel,
    parse_csv,
    get_weather,
    get_exchange_rate,
    translate,
    # P3 第三梯队只读工具（离线基础工具）
    hash_encode,
    text_diff,
    regex_test,
    unit_convert,
]
_WRITE_TOOLS = _READONLY_TOOLS + [knowledge_upload, write_file, generate_qrcode]
_ADMIN_TOOLS = _WRITE_TOOLS + [send_wecom_notification, send_email]

_TOOLS_BY_SCOPE = {
    "readonly": _READONLY_TOOLS,
    "write": _WRITE_TOOLS,
    "admin": _ADMIN_TOOLS,
}


def get_tools_for_scope(scope: str) -> list:
    """根据请求 scope 返回对应的工具列表。

    P0 修复：readonly 用户不应能通过 LLM 调用 knowledge_upload / send_email / send_wecom_notification，
    否则 HTTP 层的 require_scope("write") / require_scope("admin") 会被完全绕过。
    未知 scope 降级为最严格的 readonly。
    """
    return _TOOLS_BY_SCOPE.get(scope, _READONLY_TOOLS)
